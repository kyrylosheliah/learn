import os.path
from datetime import datetime
from matplotlib.dates import date2num
import pandas as pd
pd.set_option('display.width', 1000)
pd.set_option("display.precision", 2)
pd.set_option("display.float_format", '{:.2f}'.format)
import numpy as np
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit
from openpyxl import load_workbook

#class Object:
#    def __init__(self, **attributes):
#        self.__dict__.update(attributes)

def extract_year_fraction(date):
    # + 1 for [1;365] and [1;366]
    # + 0 for [0;364] and [0;365]
    yday = date.toordinal() - datetime(date.year, 1, 1).toordinal() # + 1
    fraction = yday / (364 + (1 if date.year % 4 == 0 else 0))
    return fraction

def f_BOD(t, BOD_0, kd):
    # Biological Oxygen Demand
    return BOD_0 * np.exp(-kd * t)
def f_DO(t, DO_sat, DO_0, BOD_0, kd, kr):
    # Dissolved Oxygen
    return DO_sat - (DO_sat-DO_0)*np.exp(-kr*t) - (kd*BOD_0/(kr-kd))*(np.exp(-kd*t)-np.exp(-kr*t))

class Model:
    def __init__(self, river_name, T, dt, date_format = "%d.%m.%Y", dpi = None):
        self.T = T
        self.dt = dt
        self.river_name = river_name
        self.date_format = date_format
        self.dpi = dpi
        print("New Streeter-Phelps model is created")
    def load_file(self):
        print(self.river_name)
        filename = self.river_name+".xls"
        if not os.path.isfile(self.river_name+".xls"):
            filename = self.river_name+".xlsx"
            if not os.path.isfile(self.river_name+".xlsx"):
                raise FileNotFoundError(self.river_name+".xls[x]")
        present_column_count = load_workbook(filename).worksheets[0].max_column
        data_columns = ['year', 'C', 'BOD', 'DO', 'v', 'h'][0:present_column_count]
        data_df = pd.read_excel(filename, sheet_name="data")
        data_df.columns = data_columns
        data_df["year"] = pd.to_datetime(data_df['year'].astype(str), format=self.date_format)
        data_df["year"] = data_df["year"].map(lambda x: x.year + extract_year_fraction(x))
        data_df.sort_values(by=['year'])
        if present_column_count == 6:
            #v = data_df["v"].mean()
            #h = data_df["h"].mean()
            v = data_df["v"].iloc[0]
            h = data_df["h"].iloc[0]
        else:
            #parameters_columns = ["v", "h"]
            parameters = pd.read_excel(filename, sheet_name="parameters")[:1]
            v = parameters.iloc[0,0]
            h = parameters.iloc[0,1]
        return (data_df, v, h)
    def calculate_kr(self):
        # calculating reaeration rate
        C = self.data_df["C"].loc[self.model_start_index]
        #C = self.data_df["C"].mean()
        #if self.h < 0.61:
        #    print("Owens-Gibbs:", end=" ")
        #    kr20 = 5.32 * np.power(self.v, 0.67) / np.power(self.h, 1.85)
        #elif self.h > 0.61 and self.h > 3.45 * np.power(self.v, 2.5):
        #    print("O'Connor-Dobbins:", end=" ")
        #    kr20 = 3.93 * np.sqrt(self.v) / np.power(self.h, 1.5)
        #else:
        #    print("Churchill:", end=" ")
        #    kr20 = 5.026 * self.v * np.power(self.h, 1.67)
        print("O'Connor-Dobbins:", end=" ")
        kr20 = 3.93 * np.sqrt(self.v) / np.power(self.h, 1.5)
        kr = kr20 * np.power(1.024, C - 20) # врахування температури
        # update attributes
        self.kr = kr
        # display results
        print("kr={:.3f}".format(kr))
    def calculate_kd(self):
        # exponential regression for BOD degradation rate
        DO_0 = self.data_df.loc[self.model_start_index, 'DO']
        BOD_0 = self.data_df.loc[self.model_start_index, 'BOD']
        y_data = self.model_df['BOD'].values
        x_data = np.arange(self.model_row_count)
        popt, pcov = curve_fit(f_BOD, x_data, y_data, bounds=([BOD_0-5,-5],[BOD_0+5,5]))
        kd = popt[1]
        # update attributes
        self.DO_0 = DO_0
        self.BOD_0 = BOD_0
        self.kd = kd
        # display results
        print("kd={:.3f}".format(popt[1]))
        plt.figure(dpi=self.dpi)
        plt.plot(x_data, y_data, 'x', color='xkcd:maroon', label = "data") # puntos
        plt.plot(x_data, f_BOD(x_data, *popt), color='xkcd:teal', label = "curve_fit") # curve_regresion
        plt.legend()
        plt.title(self.river_name + " kd exponential regression")
        plt.show()
    def model(self, dt, T):
        model_timeline = np.divide(np.arange(0,T,dt), 365) # час обчислення у роках
        DO_sat_y = self.DO_sat * np.ones(len(model_timeline))
        model_BOD = f_BOD(model_timeline, self.BOD_0, self.kd)
        model_DO = f_DO(model_timeline, self.DO_sat, self.DO_0, self.BOD_0, self.kd, self.kr)
        model_timeline = model_timeline + (self.model_start_year - self.data_start_year) + self.data_start_year
        data_BOD = self.data_df['BOD'].values
        data_DO = self.data_df['DO'].values
        data_timeline = np.array(self.data_df['year'].values)
        # update attributes
        self.model_timeline = model_timeline 
        self.DO_sat_y = DO_sat_y
        self.model_BOD = model_BOD
        self.model_DO = model_DO
        self.data_BOD = data_BOD
        self.data_DO = data_DO
        self.data_timeline = data_timeline
    def model_bod_plot(self):
        plt.figure(dpi=self.dpi)
        plt.xlabel('year')
        plt.ylabel('Biochemical Oxygen Demand')
        plt.plot(self.model_timeline, self.model_BOD, 'g-', label = "model BOD")
        plt.plot(self.data_timeline, self.data_BOD, 'go', label = "data BOD")
        plt.legend(loc='best')
        plt.title(self.river_name)
        plt.grid()
        plt.show()
    def model_do_plot(self):
        plt.figure(dpi=self.dpi)
        plt.xlabel('year')
        plt.ylabel('Dissolved Oxygen')
        plt.plot(self.model_timeline, self.model_DO, 'b-', label = "model DO")
        plt.plot(self.data_timeline, self.data_DO, 'bo', label = "data DO")
        plt.plot(self.model_timeline, self.DO_sat_y, 'c--', label = "saturated DO")
        plt.legend(loc='best')
        plt.title(self.river_name)
        plt.grid()
        plt.show()
    def fit(self, model_start_date, model_end_date, DO_sat):
        self.data_df, self.v, self.h = self.load_file()
        # initializing model from the data
        model_start_year = model_start_date.year + extract_year_fraction(model_start_date)
        model_end_year = model_end_date.year + extract_year_fraction(model_end_date)
        data_start_year = self.data_df.loc[0, 'year']
        model_start_index = self.data_df.index[self.data_df['year'] >= model_start_year].tolist()
        model_end_index = self.data_df.index[self.data_df['year'] <= model_end_year].tolist()
        if len(model_start_index) == 0 or len(model_end_index) == 0:
            raise ValueError("wrong model_start_year or model_end_year value")
        model_start_index = model_start_index[0]
        model_end_index = model_end_index[-1]
        model_df = self.data_df.loc[model_start_index:model_end_index]
        model_row_count = len(model_df.index)
        # update attributes
        self.data_start_year = data_start_year
        self.model_start_year = model_start_year
        self.DO_sat = DO_sat
        self.model_start_index = model_start_index
        self.model_df = model_df
        self.model_row_count = model_row_count
        # display results
        print(f"Selected data for model window: {self.model_row_count} entries")
        #display(model_df)
        self.calculate_kr()
        self.calculate_kd()
        self.model(self.dt, self.T)
        self.model_bod_plot()
        self.model_do_plot()
